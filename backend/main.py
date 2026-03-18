"""
Ovivo Ignition Perspective Screen Compiler — FastAPI Backend
"""

import copy
import json
import uuid
import zipfile
import tempfile
from collections import defaultdict
from pathlib import Path
from typing import Dict, List, Optional

from fastapi import FastAPI, File, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from pydantic import BaseModel

app = FastAPI(title="Ovivo Ignition Screen Compiler", version="1.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

SESSIONS: Dict[str, dict] = {}
TEMP_ROOT = Path(tempfile.gettempdir()) / "ovivo_compiler"
TEMP_ROOT.mkdir(parents=True, exist_ok=True)
PERSPECTIVE_NS = "com.inductiveautomation.perspective"


def session_dir(session_id: str) -> Path:
    d = TEMP_ROOT / session_id
    d.mkdir(parents=True, exist_ok=True)
    return d


def is_junk_entry(zip_path: str) -> bool:
    parts = zip_path.replace("\\", "/").split("/")
    return "__MACOSX" in parts or any(p == ".DS_Store" for p in parts)


def screen_page_key(screen_name: str) -> str:
    return "/" + screen_name


def view_path_str(level1: str, level2: Optional[str], screen_name: str) -> str:
    if level2:
        return f"{level1}/{level2}/{screen_name}"
    return f"{level1}/{screen_name}"


def views_zip_folder(level1: str, level2: Optional[str], screen_name: str) -> str:
    if level2:
        return f"{PERSPECTIVE_NS}/views/{level1}/{level2}/{screen_name}"
    return f"{PERSPECTIVE_NS}/views/{level1}/{screen_name}"


def read_excel_screens(excel_path: str) -> List[dict]:
    import openpyxl
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb["Page_Configuration"]
    header = next(ws.iter_rows(values_only=True))
    header_map = {str(c).strip().upper(): i for i, c in enumerate(header) if c}
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        project = row[header_map["PROJECT NAME"]] if "PROJECT NAME" in header_map else None
        folder  = row[header_map["FOLDER NAME"]]  if "FOLDER NAME"  in header_map else None
        screen  = row[header_map["SCREEN NAME"]]  if "SCREEN NAME"  in header_map else None
        if not project or not screen:
            continue
        rows.append({
            "level1":      str(project).strip(),
            "level2":      str(folder).strip() if folder else None,
            "screen_name": str(screen).strip(),
        })
    wb.close()
    return rows


def read_object_rows(excel_path: str):
    import openpyxl
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    objects, errors = [], []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        header = next(ws.iter_rows(values_only=True))
        header_map = {str(c).strip().upper(): i for i, c in enumerate(header) if c}
        required = ["SCREEN", "META", "SYMBOL", "SYSTEM"]
        if not all(x in header_map for x in required):
            continue
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            screen = row[header_map["SCREEN"]]
            meta   = row[header_map["META"]]
            symbol = row[header_map["SYMBOL"]]
            system = row[header_map["SYSTEM"]]
            if not meta:
                continue
            if not screen: errors.append(f"[{sheet}] Row {idx}: SCREEN is blank")
            if not symbol: errors.append(f"[{sheet}] Row {idx}: SYMBOL is blank")
            if not system: errors.append(f"[{sheet}] Row {idx}: SYSTEM is blank")
            if not screen or not symbol or not system:
                continue
            objects.append({"screen": str(screen).strip(), "meta": str(meta).strip(),
                            "symbol": str(symbol).strip(), "system": str(system).strip()})
    wb.close()
    return objects, errors


def read_svg_rows(excel_path: str):
    import openpyxl
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb["Page_Configuration"]
    header = next(ws.iter_rows(values_only=True))
    header_map = {str(c).strip().upper(): i for i, c in enumerate(header) if c}
    if "SVG" not in header_map:
        wb.close()
        return [], []
    rows, errors = [], []
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        svg_cell = row[header_map["SVG"]]
        screen   = row[header_map["SCREEN NAME"]]
        if not svg_cell:
            continue
        if not screen:
            errors.append(f"[Page_Configuration] Row {idx}: SCREEN NAME blank for SVG")
            continue
        for svg in str(svg_cell).split("/"):
            svg = svg.strip()
            if svg:
                rows.append({"svg": svg, "screen": str(screen).strip()})
    wb.close()
    return rows, errors


def extract_master_objects(master_view: dict) -> dict:
    return {
        obj.get("meta", {}).get("name"): obj
        for obj in master_view["root"].get("children", [])
        if obj.get("meta", {}).get("name")
    }


def extract_svg_objects(ref_svg_zip: str) -> dict:
    svg_objects = {}
    with zipfile.ZipFile(ref_svg_zip, "r") as ref:
        for item in ref.infolist():
            if not item.filename.endswith("view.json"):
                continue
            data = json.loads(ref.read(item.filename).decode())
            for obj in data.get("root", {}).get("children", []):
                meta = obj.get("meta", {}).get("name")
                if meta:
                    svg_objects[meta] = obj
    return svg_objects


def _get_ref_root(zip_path: str) -> Optional[str]:
    with zipfile.ZipFile(zip_path) as z:
        for item in z.infolist():
            if item.filename.endswith("/") and item.filename.count("/") == 1:
                return item.filename
    return None


def build_page_config_zip(rows: List[dict], ref_zip_path: str, output_path: str) -> None:
    pages = {
        screen_page_key(r["screen_name"]): {
            "title":    r["screen_name"],
            "viewPath": view_path_str(r["level1"], r["level2"], r["screen_name"]),
        }
        for r in rows
    }
    ref_root = _get_ref_root(ref_zip_path)
    with zipfile.ZipFile(ref_zip_path) as ref, \
         zipfile.ZipFile(output_path, "w", zipfile.ZIP_DEFLATED) as out:
        for item in ref.infolist():
            if is_junk_entry(item.filename):
                continue
            inner = item.filename
            if ref_root and inner.startswith(ref_root):
                inner = inner[len(ref_root):]
            if not inner:
                continue
            if item.filename.endswith("config.json"):
                out.writestr(inner, json.dumps({"pages": pages, "sharedDocks": {}}, indent=2))
            else:
                out.writestr(inner, ref.read(item.filename))


def build_views_zip(rows, objects, svg_rows, svg_objects, ref_zip_path, output_path):
    MASTER_SOURCE = f"{PERSPECTIVE_NS}/views/ovivo_library/Master_View/"
    master_files: dict = {}
    master_view = None
    ref_root = _get_ref_root(ref_zip_path)

    with zipfile.ZipFile(ref_zip_path, "r") as ref:
        for item in ref.infolist():
            if is_junk_entry(item.filename):
                continue
            inner = item.filename
            if ref_root and inner.startswith(ref_root):
                inner = inner[len(ref_root):]
            if inner.startswith(MASTER_SOURCE):
                rel = inner.replace(MASTER_SOURCE, "")
                if rel.endswith("/"):
                    continue
                data = ref.read(item.filename)
                master_files[rel] = data
                if rel.endswith("view.json"):
                    master_view = json.loads(data.decode())

    if master_view is None:
        master_view = {
            "custom": {}, "params": {},
            "props": {"defaultSize": {"height": 900, "width": 1920}},
            "root": {"children": [], "meta": {"name": "root"},
                     "props": {"style": {"classes": "ovivo_styles/display/ViewBackground"}},
                     "type": "ia.container.coord"},
        }
        master_files["view.json"] = json.dumps(master_view, indent=2).encode()
        master_files["resource.json"] = json.dumps({
            "scope": "G", "version": 1, "restricted": False, "overridable": True,
            "files": ["view.json", "thumbnail.png"],
            "attributes": {"lastModification": {"actor": "admin", "timestamp": "2026-01-01T00:00:00Z"},
                           "lastModificationSignature": ""},
        }, indent=2).encode()

    master_meta = extract_master_objects(master_view)
    views_base_inner = f"{PERSPECTIVE_NS}/views/"

    with zipfile.ZipFile(ref_zip_path, "r") as ref, \
         zipfile.ZipFile(output_path, "w", zipfile.ZIP_DEFLATED) as out:

        for item in ref.infolist():
            if is_junk_entry(item.filename):
                continue
            inner = item.filename
            if ref_root and inner.startswith(ref_root):
                inner = inner[len(ref_root):]
            if not inner:
                continue
            if inner.startswith(views_base_inner):
                continue
            out.writestr(inner, ref.read(item.filename))

        for r in rows:
            target = views_zip_folder(r["level1"], r["level2"], r["screen_name"])
            for rel_path, data in master_files.items():
                if not rel_path.endswith("view.json"):
                    out.writestr(f"{target}/{rel_path}", data)
                    continue
                view_json = copy.deepcopy(master_view)
                view_json["root"]["children"] = [
                    c for c in view_json["root"]["children"]
                    if c.get("meta", {}).get("name") not in master_meta
                ]
                children = view_json["root"]["children"]
                position_counter: dict = {}
                for obj in objects:
                    if obj["screen"] != r["screen_name"]:
                        continue
                    if obj["meta"] not in master_meta:
                        continue
                    ref_obj = master_meta[obj["meta"]]
                    base    = ref_obj["position"]
                    count   = position_counter.get(obj["meta"], 0)
                    child   = copy.deepcopy(ref_obj)
                    child["meta"]["name"] = obj["symbol"]
                    system_path = obj["system"].replace(".", "/") + "/"
                    tag = f"[{r['level1']}]{r['level1']}/{system_path}{obj['symbol']}"
                    child["props"]["params"]["tagPath"] = tag
                    child["position"] = {
                        "x": base["x"], "y": base["y"] + (count * base["height"]),
                        "width": base["width"], "height": base["height"],
                    }
                    position_counter[obj["meta"]] = count + 1
                    children.append(child)
                for s in svg_rows:
                    if s["screen"] != r["screen_name"]:
                        continue
                    if s["svg"] not in svg_objects:
                        continue
                    children.append(copy.deepcopy(svg_objects[s["svg"]]))
                out.writestr(f"{target}/{rel_path}", json.dumps(view_json, indent=2))


# ── Pydantic models ──────────────────────────────────────────────────────────

class GenerateRequest(BaseModel):
    session_id: str
    selected_screens: List[str]
    generate_page_config: bool = True
    generate_views: bool = True


# ── Routes ───────────────────────────────────────────────────────────────────

@app.post("/api/session")
def create_session():
    sid = str(uuid.uuid4())
    SESSIONS[sid] = {}
    session_dir(sid)
    return {"session_id": sid}


@app.post("/api/upload/structure")
async def upload_structure(session_id: str, file: UploadFile = File(...)):
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(400, "Structure file must be .xlsx")
    dest = session_dir(session_id) / "structure.xlsx"
    dest.write_bytes(await file.read())
    SESSIONS.setdefault(session_id, {})["structure"] = str(dest)
    return {"status": "ok", "filename": file.filename}


def _detect_ref_type(zip_path: str) -> Optional[str]:
    try:
        with zipfile.ZipFile(zip_path, "r") as z:
            names  = [n.replace("\\", "/") for n in z.namelist()]
            joined = " ".join(names)
            if "page-config/config.json" in joined:
                return "page_config"
            if "SVG" in joined.upper() and "view.json" in joined:
                return "svgs"
            if "Master_View" in joined or "view.json" in joined:
                return "views"
    except Exception:
        pass
    return None


@app.post("/api/upload/refs")
async def upload_refs(session_id: str, files: List[UploadFile] = File(...)):
    sid_dir = session_dir(session_id)
    SESSIONS.setdefault(session_id, {})
    saved = []
    for upload in files:
        raw   = await upload.read()
        fpath = sid_dir / upload.filename
        fpath.write_bytes(raw)
        detected = _detect_ref_type(str(fpath))
        if detected:
            SESSIONS[session_id][detected] = str(fpath)
        saved.append({"filename": upload.filename, "detected_as": detected or "unknown"})
    return {"status": "ok", "files": saved}


@app.post("/api/load-screens")
def load_screens(session_id: str):
    sess = SESSIONS.get(session_id)
    if not sess or "structure" not in sess:
        raise HTTPException(400, "No structure file uploaded for this session")
    try:
        screen_rows          = read_excel_screens(sess["structure"])
        objects, obj_errors  = read_object_rows(sess["structure"])
        svg_rows, svg_errors = read_svg_rows(sess["structure"])
    except Exception as e:
        raise HTTPException(500, f"Failed to parse Excel: {str(e)}")

    sess["screen_rows"] = screen_rows
    sess["objects"]     = objects
    sess["svg_rows"]    = svg_rows

    grouped: dict = defaultdict(list)
    for r in screen_rows:
        key = r["level2"] or "__ROOT__"
        grouped[key].append(r["screen_name"])

    tree = [
        {"group": (g if g != "__ROOT__" else None), "screens": s}
        for g, s in grouped.items()
    ]

    return {
        "tree":       tree,
        "total":      len(screen_rows),
        "errors":     obj_errors + svg_errors,
        "has_errors": bool(obj_errors + svg_errors),
    }


@app.post("/api/generate")
def generate(req: GenerateRequest):
    sess = SESSIONS.get(req.session_id)
    if not sess:
        raise HTTPException(400, "Invalid or expired session")
    if "screen_rows" not in sess:
        raise HTTPException(400, "Load screens first")

    selected_set = set(req.selected_screens)
    rows = [r for r in sess["screen_rows"] if r["screen_name"] in selected_set]
    if not rows:
        raise HTTPException(400, "No valid screens selected")

    out_dir = session_dir(req.session_id) / "output"
    out_dir.mkdir(exist_ok=True)
    results = {}

    if req.generate_page_config:
        if "page_config" not in sess:
            raise HTTPException(400, "No page-config reference zip uploaded")
        pc_out = str(out_dir / "generated_page_config.zip")
        build_page_config_zip(rows, sess["page_config"], pc_out)
        results["page_config"] = f"/api/download/{req.session_id}/generated_page_config.zip"

    if req.generate_views:
        if "views" not in sess:
            raise HTTPException(400, "No views reference zip uploaded")
        svg_objects = extract_svg_objects(sess["svgs"]) if "svgs" in sess else {}
        views_out = str(out_dir / "generated_views.zip")
        build_views_zip(rows, sess.get("objects", []), sess.get("svg_rows", []),
                        svg_objects, sess["views"], views_out)
        results["views"] = f"/api/download/{req.session_id}/generated_views.zip"

    return {"status": "ok", "downloads": results}


@app.get("/api/download/{session_id}/{filename}")
def download_file(session_id: str, filename: str):
    file_path = session_dir(session_id) / "output" / filename
    if not file_path.exists():
        raise HTTPException(404, "File not found")
    return FileResponse(str(file_path), filename=filename, media_type="application/zip")


@app.get("/api/health")
def health():
    return {"status": "ok"}
